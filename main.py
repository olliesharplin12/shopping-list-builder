import openpyxl
import os
from typing import List


# Recipes should be back two directories, then /Recipes/Recipes.xlsc
GRANDPARENT_DIRECTORY = os.path.abspath(os.path.join(os.getcwd(), os.pardir, os.pardir))
PATH_TO_WORKBOOK = os.path.join(GRANDPARENT_DIRECTORY, "Recipes", "Recipes.xlsx")


class Recipe:
    def __init__(self, name: str, servings: float, serving_unit: str):
        self.name = name
        self.servings = servings
        self.serving_unit = serving_unit
        self.ingredients = []

    def add_ingredient(self, name, quantity, unit):
        self.ingredients.append(Ingredient(name, quantity, unit))


class Ingredient:
    def __init__(self, name: str, quantity: float, unit: str):
        self.name = name
        self.quantity = quantity
        self.unit = unit


class RecipeList:
    def __init__(self):
        self.recipes = []

    def add_recipe(self, recipe):
        self.recipes.append(recipe)
    
    def has_recipes(self) -> bool:
        return len(self.recipes) > 0

    def contains_recipe(self, recipe) -> bool:
        for existing_recipe in self.recipes:
            if existing_recipe.name == recipe.name:
                return True
        
        return False

    def get_recipe_by_print_index(self, index) -> Recipe:
        """ Print index starts from 1 """
        if index >= 1 and index <= len(self.recipes):
            return self.recipes[index-1]
        else:
            return None

    def print_recipes(self, exclude_index=False):
        """ Print index starts from 1 """
        for i, recipe in enumerate(self.recipes):
            if exclude_index:
                print(f"{i+1}. {recipe.name}")
            else:
                print(f"- {recipe.name}")
    
    def print_unselected_recipes(self, selected_recipe_list):
        for i, recipe in enumerate(self.recipes):
            if not selected_recipe_list.contains_recipe(recipe):
                print(f"{i+1}. {recipe.name}")


def clear_terminal():
    os.system("cls" if os.name == "nt" else "clear")


def parse_recipes(path_to_workbook) -> RecipeList:
    recipes = RecipeList()

    while True:
        try:
            workbook = openpyxl.load_workbook(path_to_workbook)
            break
        except:
            input("\nClose workbook and press 'Enter' to continue.")
            clear_terminal()
    
    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]

        recipe = None
        is_header = True
        for row in worksheet.iter_rows():
            name = row[0].value
            quantity = row[1].value
            unit = row[2].value

            if is_header:
                if name != "Ingredients":
                    print(f"Skipped '{sheet}' as has no ingredients.")
                    break

                is_header = False
                recipe = Recipe(sheet, quantity, unit)
                continue
            
            if name == "Method":
                break
            elif name is None:
                continue

            recipe.add_ingredient(name, quantity, unit)
        
        if recipe is not None:
            recipes.add_recipe(recipe)
    
    return recipes


def select_recipes(all_recipes: RecipeList):
    selected_recipes = RecipeList()

    while True:
        if selected_recipes.has_recipes():
            print("\nSelected Recipes:")
            selected_recipes.print_recipes(exclude_index=True)

        print("\nAvailable Recipes:")
        all_recipes.print_unselected_recipes(selected_recipes)

        print("\nEnter the index of the recipe you wish to select or enter 'Done' to confirm the current selection:")
        user_selection = input("")

        clear_terminal()

        if user_selection.lower() == 'done':
            return selected_recipes
        
        try:
            index = int(user_selection)
        except ValueError:
            print("\nInvalid input. Ensure you are entering a number or 'Done'.")
            continue
        
        selected_recipe = all_recipes.get_recipe_by_print_index(index)
        if selected_recipe is None:
            print("\nSelected index is out of range.")
            continue
        
        if selected_recipes.contains_recipe(selected_recipe):
            print("\nThis recipe has already been selected.")
            continue
        
        selected_recipes.add_recipe(selected_recipe)


def main():
    clear_terminal()

    # Parse spreadsheet into RecipeList
    all_recipes = parse_recipes(PATH_TO_WORKBOOK)

    # Ask user for selection of recipes as RecipeList
    selected_recipes = select_recipes(all_recipes)

    # Print selected RecipeList
    clear_terminal()

    print(f"\n{len(selected_recipes.recipes)} Selected Recipes")
    for recipe in selected_recipes.recipes:
        print(f"\nRecipe: {recipe.name}")
        for ingredient in recipe.ingredients:
            print(f"Name: {ingredient.name}, Quantity: {ingredient.quantity}, Unit: {ingredient.unit}")

main()
